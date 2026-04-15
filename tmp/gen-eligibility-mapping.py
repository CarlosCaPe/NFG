"""Generate Eligibility Silver Mapping Analysis XLSX."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
mapped_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
excluded_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
pending_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
new_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
body_font = Font(name="Segoe UI", size=10)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

STATUS_FILLS = {
    "MAPPED": mapped_fill,
    "EXCLUDED": excluded_fill,
    "PENDING": pending_fill,
    "NEW_IN_SILVER": new_fill,
}

HEADERS = [
    "Entity", "MATIS Field (NewFire Model)", "Silver Column", "Type",
    "Status", "Decision", "Justification", "Scope Rule Ref", "Owner", "Resolved"
]
COL_WIDTHS = [16, 30, 32, 16, 16, 20, 55, 20, 18, 12]

# ── DATA ──
# Each row: (entity, matis_field, silver_col, dtype, status, decision, justification, scope_ref, owner, resolved)
rows = [
    # ═══════════════ ELIGIBILITY ═══════════════
    ("Eligibility", "Key", "eligibility_key", "guid (PK)", "MAPPED", "KEEP", "Primary key — surrogate from Silver SCD2", "", "Nick", "Yes"),
    ("Eligibility", "PayerId", "payer_id", "guid", "MAPPED", "KEEP", "Payer identifier — required for all scope rules", "Rules 1-6", "Nick", "Yes"),
    ("Eligibility", "PatientId", "patient_id", "guid", "MAPPED", "KEEP", "Patient identifier — case linkage", "", "Nick", "Yes"),
    ("Eligibility", "MemberId", "member_id", "varchar", "MAPPED", "KEEP", "Member identifier — primary business key for eligibility lookup", "Scope entry", "Nick", "Yes"),
    ("Eligibility", "CardId", "insurance_card_id", "varchar", "MAPPED", "KEEP", "Insurance card number — used in member search UI", "", "Nick", "Yes"),
    ("Eligibility", "MedicaidNumber", "medicaid_id", "varchar", "MAPPED", "KEEP", "Medicaid ID — dual eligibility scenarios", "DSNP #183771", "Nick", "Yes"),
    ("Eligibility", "MedicareBeneficiaryId", "medicare_beneficiary_id", "varchar", "MAPPED", "KEEP", "MBI — Medicare member identification", "", "Nick", "Yes"),
    ("Eligibility", "CoverageStart", "coverage_start_date", "date", "MAPPED", "KEEP", "Coverage effective date — eligibility window check", "Scope entry", "Nick", "Yes"),
    ("Eligibility", "CoverageEnd", "coverage_end_date", "date", "MAPPED", "KEEP", "Coverage termination date — eligibility window check", "Scope entry", "Nick", "Yes"),
    ("Eligibility", "PatientFirstName", "first_name", "varchar", "MAPPED", "KEEP", "Demographics — member search + display", "", "Nick", "Yes"),
    ("Eligibility", "PatientLastName", "last_name", "varchar", "MAPPED", "KEEP", "Demographics — member search + display", "", "Nick", "Yes"),
    ("Eligibility", "PatientMiddleInitial", "middle_name", "varchar", "MAPPED", "KEEP", "Demographics — display", "", "Nick", "Yes"),
    ("Eligibility", "PatientGender", "gender_code", "varchar", "MAPPED", "KEEP", "Demographics — clinical relevance", "", "Nick", "Yes"),
    ("Eligibility", "PatientDateOfBirth", "date_of_birth", "date", "MAPPED", "KEEP", "Demographics — age-based clinical rules", "", "Nick", "Yes"),
    ("Eligibility", "RelationshipCode", "relationship_code", "varchar", "MAPPED", "KEEP", "Subscriber vs dependent distinction", "", "Nick", "Yes"),
    ("Eligibility", "PatientPreferredLanguageCode", "preferred_language_code", "varchar", "MAPPED", "KEEP", "Language for letter generation", "", "Nick", "Yes"),
    ("Eligibility", "AccessibilityPreference", "letter_accommodation_code", "varchar", "MAPPED", "KEEP", "ADA compliance — letter format preferences", "", "Nick", "Yes"),
    ("Eligibility", "PatientHomePhone", "phone_1_type + phone_1_number", "varchar", "MAPPED", "KEEP", "Contact — home phone", "", "Nick", "Yes"),
    ("Eligibility", "PatientHomePhoneExt", "phone_1_extension", "varchar", "MAPPED", "KEEP", "Contact — extension", "", "Nick", "Yes"),
    ("Eligibility", "PatientWorkPhone", "phone_2_type + phone_2_number", "varchar", "MAPPED", "KEEP", "Contact — work phone", "", "Nick", "Yes"),
    ("Eligibility", "PatientWorkPhoneExt", "phone_2_extension", "varchar", "MAPPED", "KEEP", "Contact — extension", "", "Nick", "Yes"),
    ("Eligibility", "PatientMobilePhone", "(mapped to phone arrays)", "varchar", "MAPPED", "KEEP", "Contact — mobile", "", "Nick", "Yes"),
    ("Eligibility", "PatientMobilePhoneExt", "(mapped to phone arrays)", "varchar", "MAPPED", "KEEP", "Contact — mobile extension", "", "Nick", "Yes"),
    ("Eligibility", "StateOfPlanIssue", "state_of_issue", "varchar(2)", "MAPPED", "KEEP", "State restriction check — member state vs payer allowed set", "Rule 3", "Nick", "Yes"),
    ("Eligibility", "PlanDescription", "plan_name", "varchar", "MAPPED", "KEEP", "Display label for plan", "", "Nick", "Yes"),
    ("Eligibility", "CarrierId", "rx_carrier", "varchar", "MAPPED", "KEEP", "Cory string — carve-out identifier check", "Rule 4", "Nick", "Yes"),
    ("Eligibility", "AccountId", "rx_account", "varchar", "MAPPED", "KEEP", "Cory string — carve-out identifier check", "Rule 4", "Nick", "Yes"),
    ("Eligibility", "PharmacyGroupId", "rx_group", "varchar", "MAPPED", "KEEP", "Cory string — carve-out identifier check", "Rule 4", "Nick", "Yes"),
    ("Eligibility", "(none)", "medicare_part_c_pbp_number", "varchar", "MAPPED", "KEEP", "Medicare Part C plan benefit package number", "", "Nick", "Yes"),
    ("Eligibility", "(none)", "medicare_part_c_contract_number", "varchar", "MAPPED", "KEEP", "Medicare Part C contract", "", "Nick", "Yes"),
    ("Eligibility", "ContractNumber", "medicare_part_c_contract_number", "varchar", "MAPPED", "KEEP", "Mapped from MATIS ContractNumber", "", "Nick", "Yes"),
    ("Eligibility", "(none)", "formulary_id", "varchar", "MAPPED", "KEEP", "Drug formulary linkage", "", "Nick", "Yes"),
    ("Eligibility", "(none)", "network_id", "varchar", "MAPPED", "KEEP", "Network identifier", "", "Nick", "Yes"),
    ("Eligibility", "(none)", "plan_id", "varchar", "MAPPED", "KEEP", "Plan identifier — distinct from plan_name", "", "Nick", "Yes"),
    ("Eligibility", "CustomFields", "custom_fields", "json", "MAPPED", "KEEP", "Flexible payer-specific attributes", "", "Nick", "Yes"),
    ("Eligibility", "PatientAddressKey", "mailing_address_key", "guid (FK)", "MAPPED", "KEEP", "Join to Address — mailing", "", "Nick", "Yes"),
    ("Eligibility", "(none)", "physical_address_key", "guid (FK)", "MAPPED", "KEEP", "Join to Address — physical", "", "Nick", "Yes"),
    ("Eligibility", "(none)", "privacy_address_key", "guid (FK)", "MAPPED", "KEEP", "Join to Address — privacy/confidential", "", "Nick", "Yes"),
    ("Eligibility", "RelatedPersonKey", "member_representative_key", "guid (FK)", "MAPPED", "KEEP", "Join to RelatedPerson — representative", "", "Nick", "Yes"),
    ("Eligibility", "(none)", "subscriber_key", "guid (FK)", "MAPPED", "KEEP", "Join to RelatedPerson — subscriber", "", "Nick", "Yes"),
    ("Eligibility", "(none)", "group_key", "guid (FK)", "MAPPED", "KEEP", "Join to Group entity", "", "Nick", "Yes"),
    ("Eligibility", "BenefitPackageVersion", "benefit_package_key", "guid (FK)", "MAPPED", "KEEP", "Join to BenefitPackage entity", "", "Nick", "Yes"),
    # SCD2 fields
    ("Eligibility", "(none — SCD2 infra)", "__is_deleted", "bool", "NEW_IN_SILVER", "KEEP — Silver only", "SCD2 soft-delete flag. NOT exposed to app layer. Correct to omit from Miro.", "", "Nick", "Yes"),
    ("Eligibility", "(none — SCD2 infra)", "__start_at", "timestamp", "NEW_IN_SILVER", "KEEP — Silver only", "SCD2 version start. NOT exposed to app layer.", "", "Nick", "Yes"),
    ("Eligibility", "(none — SCD2 infra)", "__end_at", "timestamp", "NEW_IN_SILVER", "KEEP — Silver only", "SCD2 version end. NOT exposed to app layer.", "", "Nick", "Yes"),
    # ── EXCLUDED from Eligibility (orange in spreadsheet) ──
    ("Eligibility", "RelationType", "(not mapped)", "", "EXCLUDED", "DROP", "Redundant — member_representative_key & subscriber_key join to related_person directly", "", "Nick", "Yes"),
    ("Eligibility", "RelatedPersonId", "(not mapped)", "", "EXCLUDED", "DROP", "Not received in current feed; FK join sufficient", "", "Nick", "Yes"),
    ("Eligibility", "DateOfBirth (RelatedPerson)", "(not mapped)", "", "EXCLUDED", "DROP", "Not received in current feed", "", "Nick", "Yes"),
    ("Eligibility", "SSN", "(not mapped)", "", "EXCLUDED", "DROP", "PHI — not needed, not wanted. Security risk.", "", "Nick", "Yes"),
    ("Eligibility", "PatientSSN", "(not mapped)", "", "EXCLUDED", "DROP", "PHI — same as SSN. Do not store.", "", "Nick", "Yes"),
    ("Eligibility", "RelatedPersonAddrKey", "(not mapped)", "", "EXCLUDED", "DROP", "Duplicate — already on RelatedPerson.address_key", "", "Nick", "Yes"),
    ("Eligibility", "EligibilitySource", "(not mapped)", "", "EXCLUDED", "DROP", "Metadata for ETL pipeline only — not a domain field", "", "Nick", "Yes"),
    ("Eligibility", "EmployerGroupAddressKey", "(not mapped)", "", "EXCLUDED", "DROP", "Group has own table; address data not received in feed", "", "Nick", "Yes"),
    ("Eligibility", "PatientAddressVersion", "(not mapped)", "", "EXCLUDED", "DROP", "Duplicate — SCD2 handles versioning via __start_at/__end_at", "", "Nick", "Yes"),
    # ── PENDING — Nick's open questions ──
    ("Eligibility", "PayerKey", "(not mapped)", "guid", "PENDING", "NEEDS_CLARIFICATION", "Already have payer_id. Is PayerKey a different FK? If redundant, DROP.", "", "Nick + App team", "No"),
    ("Eligibility", "HasRepresentative", "(not mapped)", "bool", "PENDING", "LIKELY DROP", "Redundant — can derive from EXISTS(related_person). Boolean adds no value if FK is present.", "", "Nick", "No"),
    ("Eligibility", "LineOfBusinessVersion", "(not mapped)", "varchar", "PENDING", "NEEDS_CLARIFICATION", "Purpose unclear. If it tracks LOB config version, may be relevant; if ETL artifact, drop.", "", "Nick + Product", "No"),
    ("Eligibility", "MemberReportingId", "(not mapped)", "varchar", "PENDING", "NEEDS_CLARIFICATION", "Appears redundant with member_id + insurance_card_id. Confirm with reporting team.", "", "Nick + Reporting", "No"),
    ("Eligibility", "InsuranceGroupId", "(not mapped)", "varchar", "PENDING", "NEEDS_CLARIFICATION", "Is this different from group_key / GroupID? If same grain, drop. If payer-specific alt-key, keep.", "Rule 4?", "Nick + Config", "No"),
    ("Eligibility", "PrescriptionCarveoutType", "(not mapped)", "enum/code", "PENDING", "LIKELY KEEP", "NOT a duplicate of RxBenefitCarveOutIndicator. Indicator = bool (yes/no), Type = code (WHICH carve-out). Different data. Dropping loses granularity.", "Rule 4/6", "Nick + Product", "No"),
    ("Eligibility", "RiskPatient", "(not mapped)", "bool?", "PENDING", "NEEDS_CLARIFICATION", "Likely risk-based contract flag from MATIS. Clarify with Onco analysts — may affect case routing or reporting.", "", "Nick + Sandy/Kim", "No"),
    ("Eligibility", "CoverageTypeCode", "(not mapped)", "varchar", "PENDING", "NEEDS_CLARIFICATION", "Have Coverage Tier Cd but unused. If scope rules don't consume it, safe to drop.", "", "Nick + Config", "No"),
    ("Eligibility", "SpanType", "(not mapped)", "varchar", "PENDING", "NEEDS_CLARIFICATION", "MATIS legacy — coverage span type. Clarify meaning with analysts.", "", "Nick + MATIS analysts", "No"),
    ("Eligibility", "Name", "(not mapped)", "varchar", "PENDING", "NEEDS_CLARIFICATION", "Ambiguous — plan name? member name? Clarify source column context.", "", "Nick", "No"),
    # ── PENDING — Regulatory/Future fields ──
    ("Eligibility", "SupressLetters", "(not mapped)", "bool", "PENDING", "NEEDS_CLARIFICATION", "If app generates letters (CaseMessage), this flag is needed to suppress per-member. Product decision.", "", "Product team", "No"),
    ("Eligibility", "LetterSupressionDate", "(not mapped)", "date", "PENDING", "NEEDS_CLARIFICATION", "Companion to SupressLetters — date-bounded suppression. Same decision.", "", "Product team", "No"),
    ("Eligibility", "IsERISA", "(not mapped)", "bool", "PENDING", "LIKELY KEEP", "ERISA status affects appeal rights in post-denial pathway. If post-denial is V1, this is required.", "Post-denial", "Product + Legal", "No"),
    ("Eligibility", "IsFEHBPlan", "(not mapped)", "bool", "PENDING", "LIKELY KEEP", "Federal Employee Health Benefit — regulatory classification, similar impact as ERISA.", "Post-denial", "Product + Legal", "No"),

    # ═══════════════ RELATED PERSON ═══════════════
    ("RelatedPerson", "Key", "related_person_key", "guid (PK)", "MAPPED", "KEEP", "Primary key", "", "Nick", "Yes"),
    ("RelatedPerson", "FirstName", "first_name", "varchar", "MAPPED", "KEEP", "Representative/subscriber name", "", "Nick", "Yes"),
    ("RelatedPerson", "LastName", "last_name", "varchar", "MAPPED", "KEEP", "Representative/subscriber name", "", "Nick", "Yes"),
    ("RelatedPerson", "MiddleInitial", "middle_name", "varchar", "MAPPED", "KEEP", "Demographics", "", "Nick", "Yes"),
    ("RelatedPerson", "Gender", "gender_code", "varchar", "MAPPED", "KEEP", "Demographics", "", "Nick", "Yes"),
    ("RelatedPerson", "AddressVersion", "address_key", "guid (FK)", "MAPPED", "KEEP", "Join to Address", "", "Nick", "Yes"),
    ("RelatedPerson", "HomePhone", "phone_1_number + phone_1_type", "varchar", "MAPPED", "KEEP", "Contact", "", "Nick", "Yes"),
    ("RelatedPerson", "WorkPhone", "phone_2_number + phone_2_type", "varchar", "MAPPED", "KEEP", "Contact", "", "Nick", "Yes"),
    ("RelatedPerson", "MobilePhone", "relationship_code", "varchar", "MAPPED", "KEEP", "Relationship to member", "", "Nick", "Yes"),
    ("RelatedPerson", "(none — ETL infra)", "__insert_time", "timestamp", "NEW_IN_SILVER", "KEEP — Silver only", "Ingestion timestamp. Not exposed to app.", "", "Nick", "Yes"),

    # ═══════════════ ADDRESS ═══════════════
    ("Address", "Key", "address_key", "guid (PK)", "MAPPED", "KEEP", "Primary key", "", "Nick", "Yes"),
    ("Address", "AddressLine1", "address_line_1", "varchar", "MAPPED", "KEEP", "Street address", "", "Nick", "Yes"),
    ("Address", "AddressLine2", "address_line_2", "varchar", "MAPPED", "KEEP", "Suite/apt", "", "Nick", "Yes"),
    ("Address", "City", "city", "varchar", "MAPPED", "KEEP", "City", "", "Nick", "Yes"),
    ("Address", "State", "state", "varchar(2)", "MAPPED", "KEEP", "State code", "", "Nick", "Yes"),
    ("Address", "ZipCode", "zip_code", "varchar", "MAPPED", "KEEP", "ZIP — NOTE: typo 'zip_cope' in source Row 23, should be zip_code", "", "Nick", "Yes"),
    ("Address", "ZipPlusCode", "zip_extension", "varchar", "MAPPED", "KEEP", "ZIP+4 extension", "", "Nick", "Yes"),
    ("Address", "(none — ETL infra)", "__insert_time", "timestamp", "NEW_IN_SILVER", "KEEP — Silver only", "Ingestion timestamp. Not exposed to app.", "", "Nick", "Yes"),

    # ═══════════════ GROUP ═══════════════
    ("Group", "Key", "group_key", "guid (PK)", "MAPPED", "KEEP", "Primary key", "", "Nick", "Yes"),
    ("Group", "(none)", "organization_id", "varchar", "MAPPED", "KEEP", "Org identifier", "", "Nick", "Yes"),
    ("Group", "(none)", "organization_name", "varchar", "MAPPED", "KEEP", "Org name", "", "Nick", "Yes"),
    ("Group", "(none)", "group_id", "varchar", "MAPPED", "KEEP", "Group identifier", "", "Nick", "Yes"),
    ("Group", "Name", "group_name", "varchar", "MAPPED", "KEEP", "Group name", "", "Nick", "Yes"),
    ("Group", "(none)", "sub_group_id", "varchar", "MAPPED", "KEEP", "Sub-group identifier", "", "Nick", "Yes"),
    ("Group", "(none)", "sub_group_name", "varchar", "MAPPED", "KEEP", "Sub-group name", "", "Nick", "Yes"),
    ("Group", "ContractEffectiveDate", "contract_start_date", "date", "MAPPED", "KEEP", "Contract start", "", "Nick", "Yes"),
    ("Group", "ContractTerminationDate", "contract_end_date", "date", "MAPPED", "KEEP", "Contract end", "", "Nick", "Yes"),
    ("Group", "(none)", "custom_fields", "json", "MAPPED", "KEEP", "Flexible attributes", "", "Nick", "Yes"),
    ("Group", "(none — ETL infra)", "__insert_time", "timestamp", "NEW_IN_SILVER", "KEEP — Silver only", "Ingestion timestamp. Not exposed to app.", "", "Nick", "Yes"),
    # Excluded from Group
    ("Group", "AddressKey", "(not mapped)", "guid", "EXCLUDED", "DROP", "Group address not received in feed", "", "Nick", "Yes"),
    ("Group", "ExternalGroupId", "(not mapped)", "varchar", "EXCLUDED", "DROP", "Not mapped — purpose unclear", "", "Nick", "No"),
    ("Group", "IsErisa", "(not mapped)", "bool", "PENDING", "NEEDS_CLARIFICATION", "Duplicated on Eligibility and Group. Not received today. Resolve: which entity owns it? If needed, pick ONE location.", "Post-denial", "Nick + Product", "No"),

    # ═══════════════ BENEFIT PACKAGE ═══════════════
    ("BenefitPackage", "Key", "benefit_package_key", "guid (PK)", "MAPPED", "KEEP", "Primary key", "", "Nick", "Yes"),
    ("BenefitPackage", "MajorLineOfBusiness", "major_lob", "varchar", "MAPPED", "KEEP", "Grain hierarchy level 4 — scope validation + PDL config + case routing", "Rule 1-2", "Nick", "Yes"),
    ("BenefitPackage", "ProductType", "product_type", "varchar", "MAPPED", "KEEP", "Product classification", "Rule 2", "Nick", "Yes"),
    ("BenefitPackage", "AsoIndicator", "aso_indicator", "bool", "MAPPED", "KEEP", "ASO delegation check — ASO excluded = Out of Scope", "Rule 5", "Nick", "Yes"),
    ("BenefitPackage", "MemberBenefit (medical)", "medical_benefit_indicator", "bool", "NEW_IN_SILVER", "KEEP — ADD TO MIRO", "Scope Rule 6 requires this: medical benefit delegation check. Currently MISSING from Miro model.", "Rule 6", "Nick + App team", "No"),
    ("BenefitPackage", "MemberBenefit (rx)", "rx_benefit_indicator", "bool", "NEW_IN_SILVER", "KEEP — ADD TO MIRO", "Scope Rule 6 requires this: rx benefit delegation check. Currently MISSING from Miro model.", "Rule 6", "Nick + App team", "No"),
    ("BenefitPackage", "HasOutOfNetworkBenefit", "out_of_network_benefit_indicator", "bool", "MAPPED", "KEEP", "OON benefit — Scope Rule 6", "Rule 6", "Nick", "Yes"),
    ("BenefitPackage", "IsOutOfArea", "out_of_area_benefit_indicator", "bool", "MAPPED", "KEEP", "OOA benefit flag", "", "Nick", "Yes"),
    ("BenefitPackage", "(none)", "out_of_area_benefit_description", "varchar", "MAPPED", "KEEP", "OOA description", "", "Nick", "Yes"),
    ("BenefitPackage", "IsRxCarvedOut", "rx_benefit_carve_out_indicator", "bool", "MAPPED", "KEEP", "Rx carve-out flag — bool only (is it carved out?)", "Rule 4/6", "Nick", "Yes"),
    ("BenefitPackage", "RxCarveOutStartDate", "rx_benefit_carve_out_start_date", "date", "MAPPED", "KEEP", "Carve-out effective window start", "", "Nick", "Yes"),
    ("BenefitPackage", "RxCarveOutEndDate", "rx_benefit_carve_out_end_date", "date", "MAPPED", "KEEP", "Carve-out effective window end", "", "Nick", "Yes"),
    ("BenefitPackage", "DualEligibilityStatus", "dual_eligibility_indicator", "bool", "NEW_IN_SILVER", "KEEP — ADD TO MIRO", "DSNP dual eligibility flag — ADO #183771. Currently MISSING from Miro model.", "DSNP #183771", "Nick + Inna", "No"),
    ("BenefitPackage", "(none — ETL infra)", "__insert_time", "timestamp", "NEW_IN_SILVER", "KEEP — Silver only", "Ingestion timestamp. Not exposed to app.", "", "Nick", "Yes"),
]

# ── Build worksheet ──
ws = wb.active
ws.title = "Eligibility Mapping"

# Headers
for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Data rows
for row_idx, row_data in enumerate(rows, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin_border
    # Color by status (column 5)
    status = row_data[4]
    fill = STATUS_FILLS.get(status)
    if fill:
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

# Freeze pane
ws.freeze_panes = "A2"
# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

# ── Summary sheet ──
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 55

summary_headers = ["Category", "Count", "Detail"]
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

summary_data = [
    ("MAPPED (confirmed)", sum(1 for r in rows if r[4] == "MAPPED"), "Fields fully mapped MATIS → Silver, in Miro"),
    ("NEW_IN_SILVER (ETL infra)", sum(1 for r in rows if r[4] == "NEW_IN_SILVER" and "Silver only" in r[6]), "SCD2/ETL fields — correct to omit from Miro"),
    ("NEW_IN_SILVER (ADD TO MIRO)", sum(1 for r in rows if r[4] == "NEW_IN_SILVER" and "ADD TO MIRO" in r[6]), "medical_benefit, rx_benefit, dual_eligibility — needed by app"),
    ("EXCLUDED (confirmed DROP)", sum(1 for r in rows if r[4] == "EXCLUDED" and r[5] == "DROP"), "SSN, duplicates, ETL metadata — safe to drop"),
    ("PENDING (needs clarification)", sum(1 for r in rows if r[4] == "PENDING" and "CLARIFICATION" in r[5]), "PayerKey, InsuranceGroupId, RiskPatient, SpanType, etc."),
    ("PENDING (likely KEEP)", sum(1 for r in rows if r[4] == "PENDING" and "LIKELY KEEP" in r[5]), "PrescriptionCarveoutType, IsERISA, IsFEHBPlan"),
    ("PENDING (likely DROP)", sum(1 for r in rows if r[4] == "PENDING" and "LIKELY DROP" in r[5]), "HasRepresentative"),
    ("", "", ""),
    ("TOTAL FIELDS", len(rows), ""),
    ("RESOLVED", sum(1 for r in rows if r[9] == "Yes"), ""),
    ("OPEN", sum(1 for r in rows if r[9] == "No"), ""),
]
for row_idx, (cat, count, detail) in enumerate(summary_data, 2):
    ws2.cell(row=row_idx, column=1, value=cat).font = body_font
    ws2.cell(row=row_idx, column=2, value=count).font = body_font
    ws2.cell(row=row_idx, column=3, value=detail).font = body_font
    for col_idx in range(1, 4):
        ws2.cell(row=row_idx, column=col_idx).border = thin_border

# ── Legend sheet ──
ws3 = wb.create_sheet("Legend")
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 50
ws3.column_dimensions["C"].width = 20

legend = [
    ("Status", "Meaning", "Row Color"),
    ("MAPPED", "Field confirmed mapped from MATIS to Silver, present in Miro", "Blue"),
    ("EXCLUDED", "Field intentionally omitted — with documented reason", "Yellow"),
    ("PENDING", "Decision needed — open question from Nick or team", "Orange"),
    ("NEW_IN_SILVER", "Field exists in Silver but not in original MATIS model", "Green"),
]
for row_idx, (a, b, c) in enumerate(legend, 1):
    ws3.cell(row=row_idx, column=1, value=a).font = header_font if row_idx == 1 else body_font
    ws3.cell(row=row_idx, column=2, value=b).font = header_font if row_idx == 1 else body_font
    ws3.cell(row=row_idx, column=3, value=c).font = header_font if row_idx == 1 else body_font
    if row_idx == 1:
        for col in range(1, 4):
            ws3.cell(row=row_idx, column=col).fill = header_fill
    elif row_idx > 1:
        fill = [mapped_fill, excluded_fill, pending_fill, new_fill][row_idx - 2]
        for col in range(1, 4):
            ws3.cell(row=row_idx, column=col).fill = fill
    for col in range(1, 4):
        ws3.cell(row=row_idx, column=col).border = thin_border

# ── Save ──
out_path = r"c:\Users\CARLOS CARRILLO\git\NFG\clients\oncohealth\output\eligibility-silver-mapping-analysis.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Total rows: {len(rows)}")
print(f"Mapped: {sum(1 for r in rows if r[4] == 'MAPPED')}")
print(f"Excluded: {sum(1 for r in rows if r[4] == 'EXCLUDED')}")
print(f"Pending: {sum(1 for r in rows if r[4] == 'PENDING')}")
print(f"New in Silver: {sum(1 for r in rows if r[4] == 'NEW_IN_SILVER')}")
